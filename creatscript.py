# Code written by PV Divakar Raju and Neeraj mishra, IIT Tirupati May 2022
# This script is used to add the properties of the fiber and matrix, and the coordinates of the fiber 
# centers to the abaqus python script to model the RVE in abaqus. Here the coordinates are generated 
# Important: The coordinates are genearted using VIPER tool (https://doi.org/10.1016/J.COMPOSITESA.2016.02.026) 

import xlrd as xl
import shutil
import os
from openpyxl import load_workbook

# Begin code
f = 700 # Number of instances in the data 

for samples in range(1,f+1,1): 
    workbook    = load_workbook(" Provide appropriate path"+str(samples)+"xlsx")# coordinates excel file
    workbook2   = load_workbook("Provide excel file path")# properties excel file in same format as obtatined by datasetcreation2
    worksheet   = workbook['Sheet1']
    worksheet2  = workbook2['Sheet1']
    s           = 1;
    
    
    src = r"C:/Users/user/Desktop/New folder/new 1.py(path accordingly)"                          # ABAQUS original python script for RVE creation and applying PBC  
    dst = r"C:/Users/user/Desktop/New folder/python files/(path accordingly)"+str(samples)+".py"  # location to save all python scripts to get ready for giving input to abaqus
    shutil.copyfile(src, dst)    
    
    with open("C:/Users/user/Desktop/New folder/python files/(path accordingly)"+str(samples)+".py","r+") as f:
        contents    = f.readlines()
        text        = "mdb.models['Model-1'].parts['Part-1'].Set(cells=mdb.models['Model-1'].parts['Part-1'].cells.findAt("
        text_new    = "mdb.models['Model-1'].parts['Part-1'].setMeshControls(elemShape=WEDGE, regions=mdb.models['Model-1'].parts['Part-1'].cells.findAt(((0.0, 55.4, 0.0), ), " 
        text_new_1  = "mdb.models['Model-1'].parts['Part-1'].MaterialOrientation(additionalRotationType=ROTATION_NONE, axis=AXIS_1, fieldName='', localCsys=None, orientationType=GLOBAL, region=Region(cells=mdb.models['Model-1'].parts['Part-1'].cells.findAt(((0.0, 55.4, 0.0), ), " # material orientation (taken as global))
       
        count1 = 1
        length = worksheet.max_row
        
        for row in worksheet.iter_rows():
            val1=worksheet.cell(row=s, column=1).value
            val2=worksheet.cell(row=s, column=2).value
            if (val1<0):    # assigning sections
                val1 = 0
            if(val1>58):
                val1 = 58
            if(val2<0):
                val2 = 0
            if(val2 > 58):
                val2 = 58
            text=text+'(('+str(val1)+', '+str(val2)+', 0.0), )'
            s = s + 1
          
            if (s<=length) :
              text = text + ","
        text_new   = text_new + "), technique=SWEEP)\n"
        text_new_1 = text_new_1 + ")), stackDirection=STACK_3)"
        text       = text + ", ), name='Set-1')"
        contents.insert(111,text)
        s = 1
                
        # inserting fiber properties in the python script
        text = "mdb.models['Model-1'].materials['CF'].Elastic(type=ENGINEERING_CONSTANTS, table=(" 
        val1    =   worksheet2.cell(row=samples, column=1).value # e1
        val2    =   worksheet2.cell(row=samples, column=2).value # e2
        val3    =   worksheet2.cell(row=samples, column=3).value # nu23/(e1/e2)
        val4    =   worksheet2.cell(row=samples, column=4).value # nu12/(e1/e2)
        val5    =   worksheet2.cell(row=samples, column=5).value # nu12
        val6    =   worksheet2.cell(row=samples, column=6).value # g12
        val7    =   worksheet2.cell(row=samples, column=7).value # g23
        val8    =   worksheet2.cell(row=samples, column=8).value # nu23
        val9    =   worksheet2.cell(row=samples, column=9).value # em
        val10   =   worksheet2.cell(row=samples, column=10).value # num
        val11   =   worksheet2.cell(row=samples, column=11).value # k column (vol_fract)
        text    =   text+'('+str(val2)+', '+str(val2)+', '+str(val1)+', '+str(val8)+', '+str(val3)+', '+str(val4)+', '+str(val7)+', '+str(val6)+', '+str(val6)+')' # inserting coordinates
        text    =   text + ", ))\n"
        
        # inserting matrix properties
        text_epoxy = "    mdb.models['Model-1'].materials['EPOXY'].Elastic(table=(("+str(val9)+", "+str(val10)+"), ))\n" 
        
       # mesh convergence study
        if(val11<=0.5):
            text_k = "myAssembly.seedPartInstance(regions=(myInstance,),deviationFactor=0.1, minSizeFactor=0.1, size=0.5)"
        else:
            text_k = "myAssembly.seedPartInstance(regions=(myInstance,),deviationFactor=0.1, minSizeFactor=0.1, size=0.4)"
        contents.insert(140,text_k)
        contents.insert(104,text_epoxy)
        contents.insert(100,text)
        
        s = 1
        index=51
        for row in worksheet.iter_rows():
            val1    =   worksheet.cell(row=s, column=1).value
            val2    =   worksheet.cell(row=s, column=2).value
            val3    =   val1+2.5
            
            # inserting coordinates in the abaqus python script
            value='s1.CircleByCenterPerimeter(center=(\t'+str(val1)+'\t,\t'+str(val2)+'\t), point1=(\t'+str(val3)+'\t,\t'+str(val2)+'\t))\n' 
            xl_rows = s
                   s= s+1
            contents.insert(index+xl_rows,value)
            
    with open("C:/Users/user/Desktop/New folder/python files/(path accordingly)"+str(samples)+".py","w") as f:
        contents = "".join(map(str,contents))
        f.writelines(contents)

# End of code
